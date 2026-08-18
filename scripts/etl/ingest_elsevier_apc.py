"""Ingest Elsevier's official APC price list.

Source: https://legacyfileshare.elsevier.com/els_com_pricing/article-publishing-charge.xlsx
Verified live this session (HTTP 200, real .xlsx, no login), linked from
elsevier.com/about/policies-and-standards/pricing. Actual structure
(inspected directly): 4 metadata rows (incl. "Prices as of date: ..."),
then a header row `ISSN, Title, Business model, USD, EUR, GBP, JPY`.
Some rows carry "See journal page" instead of a number for
Subsidized-model journals - stored as apc_amount=None, not guessed.
"""
from __future__ import annotations

import datetime
import io
import re

import openpyxl
import requests

from common import etl_run, get_or_create_journal, upsert_publisher_apc
from src.db.engine import session_scope
from src.db.issn_utils import normalize_issn
from src.db.models import JournalPublisherAPC  # noqa: F401 (kept for type context)

URL = "https://legacyfileshare.elsevier.com/els_com_pricing/article-publishing-charge.xlsx"
PUBLISHER = "elsevier"
_HEADER = ("ISSN", "Title", "Business model", "USD")


def _fetch_workbook():
    resp = requests.get(URL, timeout=60)
    resp.raise_for_status()
    return openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)


def _find_header_row(ws):
    for row in ws.iter_rows(values_only=True):
        if row[:4] == _HEADER:
            return row
    return None


def _list_type(business_model: str) -> str:
    model = (business_model or "").lower()
    if "hybrid" in model:
        return "hybrid"
    if "open access" in model:
        return "fully-oa"
    return "other"


def _parse_rows(ws):
    rows_iter = ws.iter_rows(values_only=True)
    for row in rows_iter:
        if row[:4] == _HEADER:
            break
    else:
        return
    for issn, title, business_model, usd, *_rest in rows_iter:
        issn_l = normalize_issn(str(issn) if issn else "")
        if not issn_l or not title:
            continue
        apc_amount = usd if isinstance(usd, (int, float)) else None
        yield {
            "issn_l": issn_l, "title": str(title).strip(),
            "list_type": _list_type(business_model),
            "apc_amount": apc_amount, "currency": "USD",
        }


def run():
    wb = _fetch_workbook()
    ws = wb.active
    snapshot_date = datetime.datetime.utcnow()

    with etl_run(f"{PUBLISHER}_apc") as counters:
        with session_scope() as session:
            for row in _parse_rows(ws):
                counters.fetched += 1
                journal = get_or_create_journal(session, row["issn_l"], title=row["title"], publisher="Elsevier")
                upsert_publisher_apc(
                    session, journal.id, publisher=PUBLISHER, list_type=row["list_type"],
                    apc_amount=row["apc_amount"], currency=row["currency"],
                    source_url=URL, snapshot_date=snapshot_date,
                )
                counters.upserted += 1


if __name__ == "__main__":
    run()
